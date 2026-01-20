#!/usr/bin/env python3
"""Build a problems worksheet for audio verifier diffs."""
from __future__ import annotations

from dataclasses import dataclass, field

from openpyxl.utils import get_column_letter

from audio_verifier_diff import AudioVerifierDiff
from extra_audio_diff import ExtraAudioDiff
from match_audio_diff import MatchAudioDiff
from missing_audio_diff import MissingAudioDiff


@dataclass
class AudioVerifierProblemsSheetBuilder:
    headers: list[str] = field(default_factory=list)

    def __post_init__(self) -> None:
        if not self.headers:
            self.headers = ["ROLE", "type", "id", "offset", "len", "dc", "diff"]

    def build_rows(
        self,
        diffs_by_role: dict[str, list[AudioVerifierDiff]],
        role_order: list[str] | None = None,
        vetted_ids_by_role: dict[str, set[str]] | None = None,
        ignored_ids_by_role: dict[str, set[str]] | None = None,
        include_vetted: bool = False,
        include_ignored: bool = False,
    ) -> list[dict[str, object]]:
        rows: list[dict[str, object]] = []
        order = role_order if role_order is not None else sorted(diffs_by_role)
        vetted_lookup = vetted_ids_by_role or {}
        ignored_lookup = ignored_ids_by_role or {}
        if include_vetted and include_ignored:
            raise RuntimeError("Cannot include both vetted and ignored rows")
        for role in order:
            if role not in diffs_by_role:
                raise RuntimeError(f"Missing diffs for role {role}")
            vetted_ids = vetted_lookup.get(role, set())
            ignored_ids = ignored_lookup.get(role, set())
            for diff in diffs_by_role[role]:
                diff_type = self._diff_type(diff)
                if diff_type is None:
                    continue
                row = diff.to_row()
                row_id = row.get("id", "")
                is_vetted = bool(row_id) and row_id in vetted_ids
                is_ignored = bool(row_id) and row_id in ignored_ids
                if include_vetted:
                    if not is_vetted:
                        continue
                elif include_ignored:
                    if not is_ignored:
                        continue
                else:
                    if is_vetted or is_ignored:
                        continue
                rows.append(
                    {
                        "ROLE": role,
                        "type": diff_type,
                        "id": row.get("id", ""),
                        "offset": row.get("offset", ""),
                        "len": row.get("len", ""),
                        "dc": row.get("dc", ""),
                        "diff": self._diff_text(diff, row),
                    }
                )
        return rows

    def write_sheet(
        self,
        ws,
        diffs_by_role: dict[str, list[AudioVerifierDiff]],
        role_order: list[str] | None = None,
        vetted_ids_by_role: dict[str, set[str]] | None = None,
        ignored_ids_by_role: dict[str, set[str]] | None = None,
        include_vetted: bool = False,
        include_ignored: bool = False,
    ) -> None:
        ws.append(self.headers)
        rows = self.build_rows(
            diffs_by_role,
            role_order=role_order,
            vetted_ids_by_role=vetted_ids_by_role,
            ignored_ids_by_role=ignored_ids_by_role,
            include_vetted=include_vetted,
            include_ignored=include_ignored,
        )
        for row in rows:
            ws.append([row.get(header, "") for header in self.headers])
        self._format_columns(ws)

    def _diff_type(self, diff: AudioVerifierDiff) -> str | None:
        if isinstance(diff, ExtraAudioDiff):
            return "+"
        if isinstance(diff, MissingAudioDiff):
            return "-"
        if isinstance(diff, MatchAudioDiff) and diff.match_quality > 0:
            return "/"
        return None

    def _diff_text(self, diff: AudioVerifierDiff, row: dict[str, object]) -> str:
        if isinstance(diff, MatchAudioDiff):
            problem_diff = diff.problem_diff.strip()
            if problem_diff:
                return problem_diff
            return str(row.get("diff", ""))
        if isinstance(diff, MissingAudioDiff):
            expected = str(row.get("expected", "")).strip()
            return f"-{expected}" if expected else ""
        if isinstance(diff, ExtraAudioDiff):
            heard = str(row.get("heard", "")).strip()
            return f"+{heard}" if heard else ""
        return str(row.get("diff", ""))

    def _format_columns(self, ws) -> None:
        widths = {
            "ROLE": 20,
            "type": 4,
            "id": 10,
            "offset": 10,
            "len": 8,
            "dc": 5,
            "diff": 72,
        }
        for idx, header in enumerate(self.headers, start=1):
            col_letter = get_column_letter(idx)
            if header in widths:
                ws.column_dimensions[col_letter].width = widths[header]
            if header in {"offset", "len", "dc"}:
                for cell in ws[col_letter]:
                    if cell.row == 1:
                        continue
                    cell.alignment = cell.alignment.copy(horizontal="right")
            if header == "type":
                for cell in ws[col_letter]:
                    if cell.row == 1:
                        continue
                    cell.alignment = cell.alignment.copy(horizontal="center")
            if header == "diff":
                for cell in ws[col_letter]:
                    if cell.row == 1:
                        continue
                    cell.alignment = cell.alignment.copy(
                        horizontal="left",
                        vertical="top",
                        wrapText=True,
                    )
