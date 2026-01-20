#!/usr/bin/env python3
"""Build a summary worksheet for audio verifier diffs."""
from __future__ import annotations

from dataclasses import dataclass, field

from openpyxl.utils import get_column_letter

from audio_verifier_diff import AudioVerifierDiff
from extra_audio_diff import ExtraAudioDiff
from match_audio_diff import MatchAudioDiff
from missing_audio_diff import MissingAudioDiff


@dataclass
class AudioVerifierSummarySheetBuilder:
    headers: list[str] = field(default_factory=list)

    def __post_init__(self) -> None:
        if not self.headers:
            self.headers = ["role", "match", "delete", "extra", "inline_diffs", "vetted", "unvetted"]

    def build_rows(
        self,
        diffs_by_role: dict[str, list[AudioVerifierDiff]],
        role_order: list[str] | None = None,
        vetted_ids_by_role: dict[str, set[str]] | None = None,
    ) -> list[dict[str, object]]:
        rows: list[dict[str, object]] = []
        order = role_order if role_order is not None else sorted(diffs_by_role)
        vetted_lookup = vetted_ids_by_role or {}
        for role in order:
            if role not in diffs_by_role:
                raise RuntimeError(f"Missing diffs for role {role}")
            diffs = diffs_by_role[role]
            vetted_ids = vetted_lookup.get(role, set())
            match_count = sum(isinstance(diff, MatchAudioDiff) for diff in diffs)
            delete_count = 0
            extra_count = 0
            inline_diff_count = 0
            vetted_count = 0
            for diff in diffs:
                is_missing = isinstance(diff, MissingAudioDiff)
                is_extra = isinstance(diff, ExtraAudioDiff)
                is_mismatch = isinstance(diff, MatchAudioDiff) and diff.match_quality > 0
                if not (is_missing or is_extra or is_mismatch):
                    continue
                diff_id = self._diff_id(diff)
                is_vetted = bool(diff_id) and diff_id in vetted_ids
                if is_vetted:
                    vetted_count += 1
                    continue
                if is_missing:
                    delete_count += 1
                elif is_extra:
                    extra_count += 1
                elif is_mismatch:
                    inline_diff_count += 1
            rows.append(
                {
                    "role": role,
                    "match": match_count,
                    "delete": delete_count,
                    "extra": extra_count,
                    "inline_diffs": inline_diff_count,
                    "vetted": vetted_count,
                    "unvetted": delete_count + extra_count + inline_diff_count,
                }
            )
        if rows:
            totals = {
                "role": "TOTAL",
                "match": sum(row["match"] for row in rows),
                "delete": sum(row["delete"] for row in rows),
                "extra": sum(row["extra"] for row in rows),
                "inline_diffs": sum(row["inline_diffs"] for row in rows),
                "vetted": sum(row["vetted"] for row in rows),
                "unvetted": sum(row["unvetted"] for row in rows),
            }
            rows.append(totals)
        return rows

    def write_sheet(
        self,
        ws,
        diffs_by_role: dict[str, list[AudioVerifierDiff]],
        role_order: list[str] | None = None,
        vetted_ids_by_role: dict[str, set[str]] | None = None,
    ) -> None:
        ws.append(self.headers)
        rows = self.build_rows(
            diffs_by_role,
            role_order=role_order,
            vetted_ids_by_role=vetted_ids_by_role,
        )
        for row in rows:
            ws.append([row.get(header, "") for header in self.headers])
        self._format_columns(ws)

    def _format_columns(self, ws) -> None:
        widths = {
            "role": 20,
            "match": 8,
            "delete": 8,
            "extra": 8,
            "inline_diffs": 12,
            "vetted": 8,
            "unvetted": 10,
        }
        for idx, header in enumerate(self.headers, start=1):
            col_letter = get_column_letter(idx)
            if header in widths:
                ws.column_dimensions[col_letter].width = widths[header]
            if header != "role":
                for cell in ws[col_letter]:
                    if cell.row == 1:
                        continue
                    cell.alignment = cell.alignment.copy(horizontal="right")

    def _diff_id(self, diff: AudioVerifierDiff) -> str:
        if isinstance(diff, MatchAudioDiff):
            return diff.segment_id
        if isinstance(diff, MissingAudioDiff):
            return diff.segment_id
        if isinstance(diff, ExtraAudioDiff):
            return diff.extra_id
        return ""
