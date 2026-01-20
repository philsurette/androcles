#!/usr/bin/env python3
"""Write a multi-sheet audio verifier workbook with summary."""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import Workbook

from audio_verifier_diff import AudioVerifierDiff
from audio_verifier_sheet_builder import AudioVerifierSheetBuilder
from audio_verifier_summary_sheet_builder import AudioVerifierSummarySheetBuilder
from audio_verifier_problems_sheet_builder import AudioVerifierProblemsSheetBuilder


@dataclass
class AudioVerifierWorkbookWriter:
    sheet_builder: AudioVerifierSheetBuilder = field(default_factory=AudioVerifierSheetBuilder)
    summary_builder: AudioVerifierSummarySheetBuilder = field(
        default_factory=AudioVerifierSummarySheetBuilder
    )
    problems_builder: AudioVerifierProblemsSheetBuilder = field(
        default_factory=AudioVerifierProblemsSheetBuilder
    )

    def write(
        self,
        diffs_by_role: dict[str, list[AudioVerifierDiff]],
        out_path: Path,
        role_order: list[str] | None = None,
        vetted_ids_by_role: dict[str, set[str]] | None = None,
    ) -> Path:
        wb = Workbook()
        summary_ws = wb.active
        summary_ws.title = "Summary"
        self.summary_builder.write_sheet(summary_ws, diffs_by_role, role_order=role_order)

        vetted_lookup = vetted_ids_by_role or {}
        problems_ws = wb.create_sheet("Problems")
        self.problems_builder.write_sheet(
            problems_ws,
            diffs_by_role,
            role_order=role_order,
            vetted_ids_by_role=vetted_lookup,
        )
        vetted_ws = wb.create_sheet("vetted")
        self.problems_builder.write_sheet(
            vetted_ws,
            diffs_by_role,
            role_order=role_order,
            vetted_ids_by_role=vetted_lookup,
            include_vetted=True,
        )

        order = role_order if role_order is not None else sorted(diffs_by_role)
        for role in order:
            if role not in diffs_by_role:
                raise RuntimeError(f"Missing diffs for role {role}")
            ws = wb.create_sheet(self._sheet_name(role))
            self.sheet_builder.write_sheet(ws, diffs_by_role[role])

        out_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out_path)
        return out_path

    def _sheet_name(self, role: str) -> str:
        return role[:31]
