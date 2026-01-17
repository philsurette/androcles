#!/usr/bin/env python3
"""Write audio verifier diffs to an XLSX file."""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from audio_verifier_diff import AudioVerifierDiff


@dataclass
class AudioVerifierXlsxWriter:
    headers: list[str] = None

    def __post_init__(self) -> None:
        if self.headers is None:
            self.headers = [
                "Error",
                "Segment ID",
                "Offset ms",
                "Length ms",
                "Expected",
                "Heard",
                "Diff",
                "Match Quality",
            ]

    def write(self, diffs: list[AudioVerifierDiff], out_path: Path, sheet_name: str = "Verification") -> Path:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]
        ws.append(self.headers)

        for diff in diffs:
            row = diff.to_row()
            ws.append([row.get(header, "") for header in self.headers])

        self._format_columns(ws)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out_path)
        return out_path

    def _format_columns(self, ws) -> None:
        widths = {
            "Error": 4,
            "Segment ID": 10,
            "Offset ms": 10,
            "Length ms": 10,
            "Expected": 70,
            "Heard": 70,
            "Diff": 80,
            "Match Quality": 12,
        }
        for idx, header in enumerate(self.headers, start=1):
            col_letter = get_column_letter(idx)
            if header in widths:
                ws.column_dimensions[col_letter].width = widths[header]
            if header in {"Offset ms", "Length ms", "Match Quality"}:
                for cell in ws[col_letter]:
                    if cell.row == 1:
                        continue
                    cell.alignment = cell.alignment.copy(horizontal="right")
            if header == "Error":
                for cell in ws[col_letter]:
                    cell.alignment = cell.alignment.copy(horizontal="center")
