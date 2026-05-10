#!/usr/bin/env python3
"""Build a worksheet for audio verifier diffs."""
from __future__ import annotations

from dataclasses import dataclass, field

from openpyxl.utils import get_column_letter

from audio_verifier_diff import AudioVerifierDiff


@dataclass
class AudioVerifierSheetBuilder:
    headers: list[str] = field(default_factory=list)

    def __post_init__(self) -> None:
        if not self.headers:
            self.headers = [
                "status",
                "id",
                "offset",
                "len",
                "dc",
                "diff",
                "heard",
                "expected",
            ]

    def write_sheet(self, ws, diffs: list[AudioVerifierDiff]) -> None:
        ws.append(self.headers)
        for diff in diffs:
            row = diff.to_row()
            ws.append([row.get(header, "") for header in self.headers])
        self._format_columns(ws)

    def _format_columns(self, ws) -> None:
        widths = {
            "status": 4,
            "id": 10,
            "offset": 10,
            "len": 8,
            "dc": 6,
            "diff": 36,
            "heard": 36,
            "expected": 36,
        }
        for idx, header in enumerate(self.headers, start=1):
            col_letter = get_column_letter(idx)
            if header in widths:
                ws.column_dimensions[col_letter].width = widths[header]
            if header in {"offset", "len", "dc"}:
                for cell in ws[col_letter]:
                    if cell.row == 1:
                        continue
                    cell.alignment = cell.alignment.copy(horizontal="right")
            if header == "status":
                for cell in ws[col_letter]:
                    cell.alignment = cell.alignment.copy(horizontal="center")
            if header in {"diff", "heard", "expected"}:
                for cell in ws[col_letter]:
                    if cell.row == 1:
                        continue
                    cell.alignment = cell.alignment.copy(
                        horizontal="left",
                        vertical="top",
                        wrapText=True,
                    )
