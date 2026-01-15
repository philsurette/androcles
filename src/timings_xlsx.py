#!/usr/bin/env python3
"""Generate a multi-sheet XLSX timing report from build/audio/timings.csv."""
from __future__ import annotations

from collections import defaultdict

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

import paths
from segment_verifier import compute_rows


def safe_sheet_name(name: str, existing: set[str]) -> str:
    base = name[:31] if len(name) > 31 else name
    candidate = base
    idx = 1
    while candidate in existing:
        suffix = f"_{idx}"
        candidate = (base[: 31 - len(suffix)] + suffix) if len(base) + len(suffix) > 31 else base + suffix
        idx += 1
    existing.add(candidate)
    return candidate


def write_sheet(ws, headers, rows):
    ws.append(headers)
    id_col_idx = headers.index("id") + 1 if "id" in headers else None
    warn_col_idx = headers.index("warning") + 1 if "warning" in headers else None
    warn_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    missing_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    for row in rows:
        values = []
        for h in headers:
            val = row.get(h, "")
            if h in {"expected_seconds", "actual_seconds", "percent"} and val not in ("", None):
                try:
                    val = float(val)
                except ValueError:
                    pass
            values.append(val)
        ws.append(values)
        if id_col_idx and warn_col_idx:
            warn_val = values[warn_col_idx - 1]
            if warn_val == "-":
                ws.cell(row=ws.max_row, column=id_col_idx).fill = missing_fill
            elif warn_val in ("<", ">"):
                ws.cell(row=ws.max_row, column=id_col_idx).fill = warn_fill
    # Apply number formats
    for col in ws.iter_cols(min_row=2, max_row=ws.max_row):
        header = col[0].offset(row=-1).value  # header cell above data
        if header in ("expected_seconds", "actual_seconds"):
            for cell in col:
                cell.number_format = "0.0"
                cell.alignment = cell.alignment.copy(horizontal="right")
        elif header == "percent":
            for cell in col:
                cell.number_format = "0.0"
                cell.alignment = cell.alignment.copy(horizontal="right")
        elif header in ("start", "src_offset"):
            for cell in col:
                cell.alignment = cell.alignment.copy(horizontal="right")
    # Widen text column
    # Column widths
    if "warning" in headers:
        idx = headers.index("warning") + 1
        ws.column_dimensions[get_column_letter(idx)].width = 2.5  # ~7-8px
        for cell in ws[get_column_letter(idx)]:
            cell.alignment = cell.alignment.copy(horizontal="center")
    if "expected_seconds" in headers:
        idx = headers.index("expected_seconds") + 1
        ws.column_dimensions[get_column_letter(idx)].width = 5  # ~20px
    if "actual_seconds" in headers:
        idx = headers.index("actual_seconds") + 1
        ws.column_dimensions[get_column_letter(idx)].width = 5  # ~20px
    if "percent" in headers:
        idx = headers.index("percent") + 1
        ws.column_dimensions[get_column_letter(idx)].hidden = True
    if "text" in headers:
        idx = headers.index("text") + 1
        ws.column_dimensions[get_column_letter(idx)].width = 106  # ~800px
    if "role" in headers:
        idx = headers.index("role") + 1
        ws.column_dimensions[get_column_letter(idx)].width = 5  # match expected/actual columns
    if "start" in headers:
        idx = headers.index("start") + 1
        ws.column_dimensions[get_column_letter(idx)].width = 8
    if "src_offset" in headers:
        idx = headers.index("src_offset") + 1
        ws.column_dimensions[get_column_letter(idx)].width = 9
    if "id" in headers:
        idx = headers.index("id") + 1
        for cell in ws[get_column_letter(idx)]:
            cell.alignment = cell.alignment.copy(horizontal="right")


def generate_xlsx(librivox: bool = False):
    rows = compute_rows(librivox=librivox)
    headers = ["id", "warning", "expected_seconds", "actual_seconds", "percent", "start", "src_offset", "role", "text"]
    wb = Workbook()
    ws = wb.active
    ws.title = "All"
    write_sheet(ws, headers, rows)

    by_role = defaultdict(list)
    for row in rows:
        role = row.get("role", "") or "_NARRATOR"
        by_role[role].append(row)

    used_names: set[str] = {"All"}
    # Ensure narrator sheet appears immediately after "All".
    narrator_key = "_NARRATOR"
    if narrator_key in by_role:
        name = safe_sheet_name(narrator_key, used_names)
        ws_role = wb.create_sheet(title=name)
        write_sheet(ws_role, headers, by_role[narrator_key])

    for role, role_rows in sorted(by_role.items()):
        if role == narrator_key:
            continue
        name = safe_sheet_name(role, used_names)
        ws_role = wb.create_sheet(title=name)
        write_sheet(ws_role, headers, role_rows)

    out_path = paths.AUDIO_OUT_DIR / "timings.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Wrote {out_path}")

